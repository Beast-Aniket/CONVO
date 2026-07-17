#!/usr/bin/env python3
"""
ABC ID & PRN Matcher Consolidated Utility
=========================================

This application is a GUI-based utility built using Tkinter that consolidates the steps of matching student 
ABC IDs and sorting/merging records by Degree Bucket.

Working of the Code:
-------------------
1. Loads the compiled PRN -> ABC_ID mapping dictionary from multiple selected master files (.xlsx, .xls, .csv, .dbf).
2. Allows selecting multiple main target files to process.
3. Automatically identifies PRN and ABC_ID columns using fuzzy matching.
4. Allows choosing from three output modes:
   - Separate by Degree Bucket: Merges rows and splits them into separate Excel files (e.g. COMMERCE.xlsx, SCIENCE.xlsx)
     based on program/degree mapping rules.
   - Consolidate into Single File: Combines all main rows and matched ABC IDs into a single consolidated Excel file.
   - Process Individually: Matches ABC IDs in-place for each main file separately and saves copies to the output directory.
5. Interactive Mapping: When an unmapped degree/program is found, a modal dialog prompts the user to select the appropriate
   degree bucket, updates the local program_bucket_master.xlsx file, and continues execution.
"""

import os
import sys
import csv
import shutil
import threading
import queue
import re
from datetime import datetime
from pathlib import Path
from collections import OrderedDict, defaultdict
from decimal import Decimal, InvalidOperation
from typing import Any
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText

# Third-party libraries (pre-verified as installed in the environment)
try:
    import pandas as pd
    import openpyxl
    from dbfread import DBF
    import dbf
except ImportError as e:
    # Fallback to display alert before crashing if missing
    import messagebox
    messagebox.showerror("Dependency Error", f"Required library is missing: {e.name}. Please install it using pip.")
    sys.exit(1)


def clean_value(val):
    """Clean PRN/ABC ID values to prevent formatting issues and strip whitespace."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    # Standardize to string, strip whitespace
    s = str(val).strip()
    # Resolve pandas float representation (e.g., '12345.0' -> '12345')
    if s.endswith('.0'):
        s = s[:-2]
    # Remove any intermediate spaces or dashes (common in ABC IDs)
    s = s.replace(" ", "").replace("-", "")
    return s.strip()


def find_matching_columns(headers):
    """
    Fuzzy matches headers to identify PRN and ABC ID columns.
    Returns a tuple of (prn_column_name, abc_column_name).
    """
    prn_col = None
    abc_col = None

    # Normalized comparison lookup lists
    prn_terms = ["prn", "prnno", "prnnum", "prnnumber", "permanentregistrationno", "permanentregistrationnumber", "enrollmentno", "enrollmentnumber", "seatno", "seatnumber"]
    abc_terms = ["abcid", "abcidno", "abcidnumber", "abc_id", "abc", "academicbankofcreditsid", "academicbankofcreditid", "academicbankofcredits", "abcno", "abcadd"]

    # Exact matching (ignoring spaces, case, underscores, dashes)
    for h in headers:
        if h is None:
            continue
        h_clean = str(h).lower().replace(" ", "").replace("_", "").replace("-", "").strip()
        
        if h_clean in prn_terms and prn_col is None:
            prn_col = h
        if h_clean in abc_terms and abc_col is None:
            abc_col = h

    # Partial/Fuzzy matching if exact match not found
    if prn_col is None:
        for h in headers:
            if h is None:
                continue
            h_clean = str(h).lower()
            if "prn" in h_clean:
                prn_col = h
                break

    if abc_col is None:
        for h in headers:
            if h is None:
                continue
            h_clean = str(h).lower()
            if "abc" in h_clean:
                abc_col = h
                break

    return prn_col, abc_col


# ==========================================
# Degree Bucket Mapping & Separation Helpers
# ==========================================

STANDARD_COLUMNS = [
    "LotNo", "Conv ID", "Faculty", "PRNERN", "ProgType", "APPL_NO", "SEAT_NO",
    "COLL_NO", "COLL_NAME", "COLL_NAMEM", "StudLastName", "StudFirstName",
    "StudMidddleName", "StudMotherName", "NAME", "NAME_MARAT", "SEX", "ABBR",
    "CLASS", "MCLASS", "SUB1", "SUB1_NAME", "SUB1_NAMEM", "SUB2", "SUB2_NAME",
    "SUB2_NAMEM", "DEGNM", "MDEGNM", "SUBDEGNM", "MSUBDEGNM", "PER", "MPER",
]

TEXT_ONLY_COLUMNS = {"PRNERN", "MCLASS"}
CLASS_COLUMN = "CLASS"
DEFAULT_SHEET_NAME = "Structured"
DEFAULT_BUCKET_MASTER = "program_bucket_master.xlsx"
SHEET_NOTES = "Read_Me"
SHEET_BUCKETS = "Bucket_List"
SHEET_DEGREES = "Degree_List"
SHEET_PROGRAMS = "Program_Mapping"

DEFAULT_BUCKETS = [
    ("COMMERCE", "convocation_commerce@mu.ac.in"),
    ("MANAGEMENT", "convocation_management@mu.ac.in"),
    ("SCIENCE", "convocation_Science@mu.ac.in"),
    ("ENGINEERING", "convocation_engineering@mu.ac.in"),
    ("ART", "convocation_art@mu.ac.in"),
    ("LAW", "convocation_law@mu.ac.in"),
]

PROGRAM_MASTER_COLUMNS = [
    "PROG_NO", "ABBR", "FACULTY", "DEGNM", "MDEGNM", "SUBDEGNM",
    "MSUBDEGNM", "MERGE_NAME", "EMAIL", "SOURCE", "STATUS",
]

DEGREE_MASTER_COLUMNS = [
    "DEGNM", "MERGE_NAME", "EMAIL", "SOURCE", "STATUS", "PROGRAM_COUNT", "FACULTIES",
]


class MissingMappingError(Exception):
    def __init__(self, row_dict: dict[str, Any], file_name: str):
        self.row_dict = row_dict
        self.file_name = file_name


def normalize_header(value: Any) -> str:
    if value is None: return ""
    return str(value).strip()


def normalize_lookup_key(value: Any) -> str:
    if value is None: return ""
    text = str(value).strip()
    if not text: return ""
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def is_blank_row(values: tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def as_text(value: Any) -> str | None:
    if value is None: return None
    if isinstance(value, str): return value
    if isinstance(value, bool): return "TRUE" if value else "FALSE"
    if isinstance(value, int): return str(value)
    if isinstance(value, float):
        if value.is_integer(): return str(int(value))
        text = format(value, ".15g")
        if "e" in text.lower():
            return format(Decimal(str(value)), "f").rstrip("0").rstrip(".")
        return text
    return str(value)


def parse_numeric_class(value: Any) -> float | None:
    if value is None or isinstance(value, bool): return None
    if isinstance(value, (int, float)): return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text: return None
        try:
            return float(Decimal(text))
        except InvalidOperation:
            return None
    return None


def safe_filename(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\\\|?*]+', "_", name.strip())
    cleaned = re.sub(r"\s+", "_", cleaned)
    cleaned = re.sub(r"_+", "_", cleaned).strip("._")
    return cleaned or "OUTPUT"


def classify_degree_name(degree_name: Any) -> tuple[str | None, str]:
    degree_key = normalize_lookup_key(degree_name)
    if not degree_key: return None, "EMPTY_DEGREE"

    if re.search(r"\bLAWS?\b", degree_key): return "LAW", "AUTO:LAW"
    if "BUSINESS ADMINISTRATION" in degree_key or re.search(r"\bMANAGEMENT\b", degree_key):
        return "MANAGEMENT", "AUTO:MANAGEMENT"
    if re.search(r"\bCOMMERCE\b", degree_key): return "COMMERCE", "AUTO:COMMERCE"
    if re.search(r"\bARCHITECTURE\b", degree_key): return "SCIENCE", "AUTO:ARCHITECTURE_TO_SCIENCE"
    if re.search(r"\bENGINEERING\b", degree_key): return "ENGINEERING", "AUTO:ENGINEERING"
    if "COMPUTER APPLICATIONS" in degree_key or re.search(r"\bSCIENCE\b|\bPHARMACY\b", degree_key):
        return "SCIENCE", "AUTO:SCIENCE"
    if ("FINE ARTS" in degree_key or "PHYSICAL EDUCATION" in degree_key or 
        "SOCIAL WORK" in degree_key or "EDUCATION" in degree_key or 
        "MUSIC" in degree_key or re.search(r"\bARTS?\b", degree_key)):
        return "ART", "AUTO:ART"
    return None, "REVIEW_NEEDED"


def create_empty_master() -> dict[str, Any]:
    master = {
        "buckets": OrderedDict(),
        "degree_rows": OrderedDict(),
        "program_rows": OrderedDict(),
        "program_master_path": "",
        "indexes": {"degrees": {}, "programs_by_prog": {}, "programs_by_abbr": {}},
    }
    for bucket_name, email in DEFAULT_BUCKETS:
        add_bucket(master, bucket_name, email)
    return master


def add_bucket(master: dict[str, Any], bucket_name: Any, email: Any = "") -> str:
    display_name = str(bucket_name).strip()
    if not display_name: return ""
    key = normalize_lookup_key(display_name)
    existing = master["buckets"].get(key)
    if existing:
        if str(email).strip(): existing["email"] = str(email).strip()
        return existing["name"]
    master["buckets"][key] = {"name": display_name, "email": str(email).strip()}
    return display_name


def canonical_bucket_name(master: dict[str, Any], bucket_name: Any) -> str:
    display_name = str(bucket_name).strip() if bucket_name is not None else ""
    if not display_name: return ""
    key = normalize_lookup_key(display_name)
    existing = master["buckets"].get(key)
    if existing: return existing["name"]
    return add_bucket(master, display_name, "")


def get_bucket_email(master: dict[str, Any], bucket_name: Any) -> str:
    key = normalize_lookup_key(bucket_name)
    if not key: return ""
    bucket = master["buckets"].get(key)
    return bucket["email"] if bucket else ""


def ordered_union(values: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        if value in seen: continue
        seen.add(value)
        ordered.append(value)
    return ordered


def get_program_row_key(program_no: Any, abbr: Any, row_number: int | None = None) -> str:
    program_key = normalize_lookup_key(program_no)
    if program_key: return f"PROG::{program_key}"
    abbr_key = normalize_lookup_key(abbr)
    if abbr_key: return f"ABBR::{abbr_key}"
    if row_number is not None: return f"ROW::{row_number}"
    return "ROW::UNKNOWN"


def rebuild_master_indexes(master: dict[str, Any]) -> None:
    degree_index: dict[str, dict[str, Any]] = {}
    programs_by_prog: dict[str, dict[str, Any]] = {}
    programs_by_abbr: dict[str, dict[str, Any]] = {}

    for degree_key, row in master["degree_rows"].items():
        degree_index[degree_key] = row
        bucket_name = canonical_bucket_name(master, row.get("MERGE_NAME"))
        row["MERGE_NAME"] = bucket_name
        row["EMAIL"] = get_bucket_email(master, bucket_name) if bucket_name else ""

    for row_key, row in master["program_rows"].items():
        bucket_name = canonical_bucket_name(master, row.get("MERGE_NAME"))
        row["MERGE_NAME"] = bucket_name
        row["EMAIL"] = get_bucket_email(master, bucket_name) if bucket_name else ""

        program_key = normalize_lookup_key(row.get("PROG_NO"))
        abbr_key = normalize_lookup_key(row.get("ABBR"))
        if program_key: programs_by_prog[program_key] = row
        if abbr_key: programs_by_abbr[abbr_key] = row

    master["indexes"] = {
        "degrees": degree_index,
        "programs_by_prog": programs_by_prog,
        "programs_by_abbr": programs_by_abbr,
    }


def read_sheet_dicts(worksheet) -> list[dict[str, Any]]:
    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None: return []
    headers = [normalize_header(value) for value in header_row]
    records: list[dict[str, Any]] = []
    for row in rows:
        if row is None or is_blank_row(row): continue
        record: dict[str, Any] = {}
        for index in range(min(len(headers), len(row))):
            header = headers[index]
            if not header: continue
            record[header] = row[index]
        if record: records.append(record)
    return records


def load_bucket_master(bucket_master_path: Path) -> dict[str, Any]:
    if not bucket_master_path.exists(): return create_empty_master()
    workbook = openpyxl.load_workbook(bucket_master_path, read_only=True, data_only=True)
    master = create_empty_master()
    try:
        if SHEET_BUCKETS in workbook.sheetnames:
            for row in read_sheet_dicts(workbook[SHEET_BUCKETS]):
                add_bucket(master, row.get("MERGE_NAME"), row.get("EMAIL"))

        if SHEET_DEGREES in workbook.sheetnames:
            for row in read_sheet_dicts(workbook[SHEET_DEGREES]):
                degree_name = str(row.get("DEGNM") or "").strip()
                if not degree_name: continue
                bucket_name = canonical_bucket_name(master, row.get("MERGE_NAME"))
                degree_key = normalize_lookup_key(degree_name)
                master["degree_rows"][degree_key] = {
                    "DEGNM": degree_name,
                    "MERGE_NAME": bucket_name,
                    "EMAIL": get_bucket_email(master, bucket_name) if bucket_name else "",
                    "SOURCE": str(row.get("SOURCE") or "").strip(),
                    "STATUS": str(row.get("STATUS") or "").strip(),
                    "PROGRAM_COUNT": row.get("PROGRAM_COUNT"),
                    "FACULTIES": str(row.get("FACULTIES") or "").strip(),
                }

        if SHEET_PROGRAMS in workbook.sheetnames:
            for row_number, row in enumerate(read_sheet_dicts(workbook[SHEET_PROGRAMS]), start=2):
                bucket_name = canonical_bucket_name(master, row.get("MERGE_NAME"))
                row_key = get_program_row_key(row.get("PROG_NO"), row.get("ABBR"), row_number=row_number)
                master["program_rows"][row_key] = {
                    "PROG_NO": str(row.get("PROG_NO") or "").strip(),
                    "ABBR": str(row.get("ABBR") or "").strip(),
                    "FACULTY": str(row.get("FACULTY") or "").strip(),
                    "DEGNM": str(row.get("DEGNM") or "").strip(),
                    "MDEGNM": str(row.get("MDEGNM") or "").strip(),
                    "SUBDEGNM": str(row.get("SUBDEGNM") or "").strip(),
                    "MSUBDEGNM": str(row.get("MSUBDEGNM") or "").strip(),
                    "MERGE_NAME": bucket_name,
                    "EMAIL": get_bucket_email(master, bucket_name) if bucket_name else "",
                    "SOURCE": str(row.get("SOURCE") or "").strip(),
                    "STATUS": str(row.get("STATUS") or "").strip(),
                }
    finally:
        workbook.close()
    rebuild_master_indexes(master)
    return master


def save_bucket_master(bucket_master_path: Path, master: dict[str, Any]) -> None:
    bucket_master_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    notes = workbook.active
    notes.title = SHEET_NOTES
    notes.append(["What to edit"])
    notes.append(["Edit MERGE_NAME in Degree_List for whole-degree defaults. Edit MERGE_NAME in Program_Mapping for program-specific overrides."])
    notes.append(["The merge script reads this Excel file on every run. Your edits here are used for future separation."])
    notes.append(["Bucket_List controls the available bucket names and emails shown in the GUI."])

    bucket_sheet = workbook.create_sheet(SHEET_BUCKETS)
    bucket_sheet.append(["MERGE_NAME", "EMAIL"])
    for bucket in master["buckets"].values():
        bucket_sheet.append([bucket["name"], bucket["email"]])

    degree_sheet = workbook.create_sheet(SHEET_DEGREES)
    degree_sheet.append(DEGREE_MASTER_COLUMNS)
    for row in master["degree_rows"].values():
        degree_sheet.append([
            row.get("DEGNM", ""), row.get("MERGE_NAME", ""), row.get("EMAIL", ""),
            row.get("SOURCE", ""), row.get("STATUS", ""), row.get("PROGRAM_COUNT", ""),
            row.get("FACULTIES", "")
        ])

    program_sheet = workbook.create_sheet(SHEET_PROGRAMS)
    program_sheet.append(PROGRAM_MASTER_COLUMNS)
    for row in master["program_rows"].values():
        program_sheet.append([
            row.get("PROG_NO", ""), row.get("ABBR", ""), row.get("FACULTY", ""),
            row.get("DEGNM", ""), row.get("MDEGNM", ""), row.get("SUBDEGNM", ""),
            row.get("MSUBDEGNM", ""), row.get("MERGE_NAME", ""), row.get("EMAIL", ""),
            row.get("SOURCE", ""), row.get("STATUS", "")
        ])
    workbook.save(bucket_master_path)


def read_program_master_records(program_master_path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(program_master_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None: return []

        headers = [normalize_header(value) for value in header_row]
        records: list[dict[str, str]] = []
        for row in rows:
            if row is None or is_blank_row(row): continue
            record: dict[str, str] = {}
            for index in range(min(len(headers), len(row))):
                header = headers[index]
                if not header: continue
                record[header] = "" if row[index] is None else str(row[index]).strip()
            if record: records.append(record)
        return records
    finally:
        workbook.close()


def resolve_row_from_indexes(master: dict[str, Any], program_no: Any = None, abbr: Any = None, degree_name: Any = None) -> tuple[str, dict[str, Any] | None]:
    abbr_key = normalize_lookup_key(abbr)
    if abbr_key:
        row = master["indexes"]["programs_by_abbr"].get(abbr_key)
        if row and str(row.get("MERGE_NAME") or "").strip(): return row["MERGE_NAME"], row

    program_key = normalize_lookup_key(program_no)
    if program_key:
        row = master["indexes"]["programs_by_prog"].get(program_key)
        if row and str(row.get("MERGE_NAME") or "").strip(): return row["MERGE_NAME"], row

    degree_key = normalize_lookup_key(degree_name)
    if degree_key:
        row = master["indexes"]["degrees"].get(degree_key)
        if row and str(row.get("MERGE_NAME") or "").strip(): return row["MERGE_NAME"], row

    return "", None


def sync_master_from_program_master(master: dict[str, Any], program_master_path: Path) -> None:
    records = read_program_master_records(program_master_path)
    existing_degree_rows = master["degree_rows"]
    existing_program_rows = master["program_rows"]
    updated_program_rows: OrderedDict[str, dict[str, Any]] = OrderedDict()
    degree_stats: dict[str, dict[str, Any]] = defaultdict(lambda: {"count": 0, "faculties": set(), "sample_degree": ""})

    for record in records:
        degree_name = str(record.get("DEGNM") or "").strip()
        degree_key = normalize_lookup_key(degree_name)
        program_no = record.get("PROG_NO")
        abbr = record.get("ABBR")
        row_key = get_program_row_key(program_no, abbr)
        existing_program = existing_program_rows.get(row_key)
        existing_degree = existing_degree_rows.get(degree_key)

        bucket_name = ""
        source = ""
        if existing_program and str(existing_program.get("MERGE_NAME") or "").strip():
            bucket_name = canonical_bucket_name(master, existing_program.get("MERGE_NAME"))
            source = str(existing_program.get("SOURCE") or "").strip() or "EXCEL:PROGRAM"
        elif existing_degree and str(existing_degree.get("MERGE_NAME") or "").strip():
            bucket_name = canonical_bucket_name(master, existing_degree.get("MERGE_NAME"))
            source = str(existing_degree.get("SOURCE") or "").strip() or "EXCEL:DEGREE"
        else:
            bucket_name, source = classify_degree_name(degree_name)
            bucket_name = canonical_bucket_name(master, bucket_name)

        status = "MAPPED" if bucket_name else "REVIEW_NEEDED"
        updated_program_rows[row_key] = {
            "PROG_NO": str(program_no or "").strip(), "ABBR": str(abbr or "").strip(),
            "FACULTY": str(record.get("FACULTY") or "").strip(), "DEGNM": degree_name,
            "MDEGNM": str(record.get("MDEGNM") or "").strip(), "SUBDEGNM": str(record.get("SUBDEGNM") or "").strip(),
            "MSUBDEGNM": str(record.get("MSUBDEGNM") or "").strip(), "MERGE_NAME": bucket_name,
            "EMAIL": get_bucket_email(master, bucket_name) if bucket_name else "",
            "SOURCE": source, "STATUS": status,
        }
        degree_stats[degree_key]["count"] += 1
        faculty = str(record.get("FACULTY") or "").strip()
        if faculty: degree_stats[degree_key]["faculties"].add(faculty)
        degree_stats[degree_key]["sample_degree"] = degree_name

    for row_key, existing_row in existing_program_rows.items():
        if row_key in updated_program_rows: continue
        bucket_name = canonical_bucket_name(master, existing_row.get("MERGE_NAME"))
        updated_program_rows[row_key] = {
            **existing_row, "MERGE_NAME": bucket_name,
            "EMAIL": get_bucket_email(master, bucket_name) if bucket_name else "",
            "STATUS": existing_row.get("STATUS") or ("MAPPED" if bucket_name else "REVIEW_NEEDED"),
        }
        degree_name = str(existing_row.get("DEGNM") or "").strip()
        degree_key = normalize_lookup_key(degree_name)
        if degree_name: degree_stats[degree_key]["sample_degree"] = degree_name

    master["program_rows"] = updated_program_rows
    degree_keys = ordered_union(list(degree_stats.keys()) + list(existing_degree_rows.keys()))
    updated_degree_rows: OrderedDict[str, dict[str, Any]] = OrderedDict()

    for degree_key in degree_keys:
        stats = degree_stats.get(degree_key, {"count": 0, "faculties": set(), "sample_degree": ""})
        existing_degree = existing_degree_rows.get(degree_key)
        degree_name = str(existing_degree.get("DEGNM") or "").strip() if existing_degree else stats["sample_degree"]
        bucket_name = ""
        source = ""
        if existing_degree and str(existing_degree.get("MERGE_NAME") or "").strip():
            bucket_name = canonical_bucket_name(master, existing_degree.get("MERGE_NAME"))
            source = str(existing_degree.get("SOURCE") or "").strip() or "EXCEL:DEGREE"
        else:
            bucket_name, source = classify_degree_name(degree_name)
            bucket_name = canonical_bucket_name(master, bucket_name)

        status = "MAPPED" if bucket_name else "REVIEW_NEEDED"
        updated_degree_rows[degree_key] = {
            "DEGNM": degree_name, "MERGE_NAME": bucket_name,
            "EMAIL": get_bucket_email(master, bucket_name) if bucket_name else "",
            "SOURCE": source, "STATUS": status, "PROGRAM_COUNT": stats["count"],
            "FACULTIES": ", ".join(sorted(stats["faculties"])),
        }
    master["degree_rows"] = updated_degree_rows
    master["program_master_path"] = str(program_master_path.resolve())
    rebuild_master_indexes(master)


def can_promote_to_degree_default(master: dict[str, Any], degree_key: str, chosen_bucket: str) -> bool:
    existing_degree = master["degree_rows"].get(degree_key)
    existing_bucket = canonical_bucket_name(master, existing_degree.get("MERGE_NAME")) if existing_degree else ""
    if existing_bucket and existing_bucket != chosen_bucket: return False

    mapped_buckets = set()
    for row in master["program_rows"].values():
        if normalize_lookup_key(row.get("DEGNM")) != degree_key: continue
        bucket_name = canonical_bucket_name(master, row.get("MERGE_NAME"))
        if bucket_name: mapped_buckets.add(bucket_name)

    if not mapped_buckets: return True
    return mapped_buckets == {chosen_bucket}


def remember_manual_selection(master: dict[str, Any], chosen_bucket: str, row_dict: dict[str, Any]) -> None:
    bucket_name = canonical_bucket_name(master, chosen_bucket)
    program_no = row_dict.get("APPL_NO") or row_dict.get("PROG_NO")
    abbr = row_dict.get("ABBR")

    matched_row = None
    abbr_key = normalize_lookup_key(abbr)
    if abbr_key: matched_row = master["indexes"]["programs_by_abbr"].get(abbr_key)

    if matched_row is None:
        program_key = normalize_lookup_key(program_no)
        if program_key: matched_row = master["indexes"]["programs_by_prog"].get(program_key)

    if matched_row is None:
        row_key = get_program_row_key(program_no, abbr, row_number=len(master["program_rows"]) + 2)
        matched_row = {
            "PROG_NO": str(program_no or "").strip(), "ABBR": str(abbr or "").strip(),
            "FACULTY": str(row_dict.get("Faculty") or row_dict.get("FACULTY") or "").strip(),
            "DEGNM": str(row_dict.get("DEGNM") or "").strip(), "MDEGNM": str(row_dict.get("MDEGNM") or "").strip(),
            "SUBDEGNM": str(row_dict.get("SUBDEGNM") or "").strip(), "MSUBDEGNM": str(row_dict.get("MSUBDEGNM") or "").strip(),
            "MERGE_NAME": "", "EMAIL": "", "SOURCE": "", "STATUS": "",
        }
        master["program_rows"][row_key] = matched_row

    matched_row["MERGE_NAME"] = bucket_name
    matched_row["EMAIL"] = get_bucket_email(master, bucket_name)
    matched_row["SOURCE"] = "MANUAL:GUI_SELECTION"
    matched_row["STATUS"] = "MAPPED"

    degree_name = str(row_dict.get("DEGNM") or "").strip()
    degree_key = normalize_lookup_key(degree_name)
    if degree_name and can_promote_to_degree_default(master, degree_key, bucket_name):
        degree_row = master["degree_rows"].get(degree_key)
        if degree_row is None:
            degree_row = {
                "DEGNM": degree_name, "MERGE_NAME": "", "EMAIL": "", "SOURCE": "", "STATUS": "",
                "PROGRAM_COUNT": "", "FACULTIES": "",
            }
            master["degree_rows"][degree_key] = degree_row
        degree_row["MERGE_NAME"] = bucket_name
        degree_row["EMAIL"] = get_bucket_email(master, bucket_name)
        degree_row["SOURCE"] = "MANUAL:GUI_SELECTION"
        degree_row["STATUS"] = "MAPPED"

    rebuild_master_indexes(master)


def read_rows_from_file(file_path):
    ext = os.path.splitext(file_path.lower())[1]
    rows = []
    headers = []
    if ext == '.dbf':
        db = DBF(file_path, ignore_missing_memofile=True)
        headers = list(db.field_names)
        for r in db:
            rows.append(dict(r))
    elif ext in ['.xlsx', '.xls']:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        # Find headers
        header_row_idx = 1
        for r in range(1, min(15, ws.max_row + 1)):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            temp_headers = [str(v).strip() if v is not None else "" for v in row_vals]
            p_col, _ = find_matching_columns(temp_headers)
            if p_col is not None:
                header_row_idx = r
                headers = [h for h in temp_headers if h]
                break
        if not headers:
            headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
            headers = [h for h in headers if h]
            header_row_idx = 1
            
        for r in range(header_row_idx + 1, ws.max_row + 1):
            row_dict = {}
            for col_idx, h in enumerate(headers, 1):
                row_dict[h] = ws.cell(row=r, column=col_idx).value
            if not is_blank_row(tuple(row_dict.values())):
                rows.append(row_dict)
        wb.close()
    elif ext == '.csv':
        # Sniff delimiter and encoding
        encoding = 'utf-8-sig'
        delimiter = ','
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                sample = f.read(4096)
                f.seek(0)
                if sample:
                    try:
                        dialect = csv.Sniffer().sniff(sample)
                        delimiter = dialect.delimiter
                    except Exception:
                        pass
        except Exception:
            encoding = 'cp1252'
            
        with open(file_path, 'r', encoding=encoding, newline='') as f:
            reader = csv.reader(f, delimiter=delimiter)
            all_lines = list(reader)
        if all_lines:
            header_row_idx = 0
            for idx in range(min(15, len(all_lines))):
                temp_headers = [str(h).strip() for h in all_lines[idx]]
                p_col, _ = find_matching_columns(temp_headers)
                if p_col is not None:
                    header_row_idx = idx
                    headers = [h for h in temp_headers if h]
                    break
            if not headers:
                headers = [str(h).strip() for h in all_lines[0]]
                headers = [h for h in headers if h]
                header_row_idx = 0
                
            for r_idx in range(header_row_idx + 1, len(all_lines)):
                line = all_lines[r_idx]
                row_dict = {}
                for col_idx, h in enumerate(headers):
                    if col_idx < len(line):
                        row_dict[h] = line[col_idx]
                    else:
                        row_dict[h] = ""
                if not is_blank_row(tuple(row_dict.values())):
                    rows.append(row_dict)
    return headers, rows


def write_bucket_workbook(output_path: Path, sheet_name: str, column_order: list[str], rows: list[dict[str, Any]]) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(column_order)

    for row_index, row_dict in enumerate(rows, start=2):
        for col_index, column_name in enumerate(column_order, start=1):
            value = row_dict.get(column_name)
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.value = value

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


# ==========================================
# Core Application Class
# ==========================================

class ABCIDMatcherApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ABC ID & PRN Matcher")
        self.root.geometry("850x780")
        self.root.minsize(750, 700)
        self.root.configure(bg="#0F172A") # Slate 900

        # State Variables
        self.main_file_path = ""
        self.main_files_paths = []
        self.master_files_paths = []
        self.output_dir_path = ""
        self.output_mode = tk.StringVar(value="separate_degree")
        self.msg_queue = queue.Queue()

        # Custom ttk styles
        self.setup_styles()

        # Build Interface
        self.create_widgets()

        # Start message queue poller
        self.root.after(100, self.process_queue)

    def setup_styles(self):
        self.style = ttk.Style()
        self.style.theme_use('default')
        # Progress Bar Styling
        self.style.configure(
            "Indigo.Horizontal.TProgressbar",
            troughcolor="#1E293B",
            background="#6366F1",
            thickness=12,
            borderwidth=0
        )

    def make_flat_button(self, parent, text, command, bg="#4F46E5", hover_bg="#4338CA", fg="white", font=("Segoe UI", 9, "bold")):
        """Utility helper to create beautiful flat Tkinter buttons with hover effects."""
        btn = tk.Button(
            parent,
            text=text,
            command=command,
            bg=bg,
            fg=fg,
            activebackground=hover_bg,
            activeforeground=fg,
            font=font,
            relief="flat",
            bd=0,
            padx=14,
            pady=6,
            cursor="hand2"
        )
        def on_enter(e):
            if btn['state'] != tk.DISABLED:
                btn.config(bg=hover_bg)
        def on_leave(e):
            if btn['state'] != tk.DISABLED:
                btn.config(bg=bg)
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        return btn

    def create_widgets(self):
        # 1. Header Banner
        header_frame = tk.Frame(self.root, bg="#1E293B", padx=20, pady=15)
        header_frame.pack(fill="x", side="top")
        
        title_label = tk.Label(
            header_frame, 
            text="ABC ID & PRN Matcher", 
            font=("Segoe UI", 16, "bold"), 
            fg="#818CF8", 
            bg="#1E293B"
        )
        title_label.pack(anchor="w")
        
        subtitle_label = tk.Label(
            header_frame, 
            text="Securely match and populate student ABC ID values based on PRN and optionally sort/merge by Degree Bucket.", 
            font=("Segoe UI", 9), 
            fg="#94A3B8", 
            bg="#1E293B"
        )
        subtitle_label.pack(anchor="w", pady=(2, 0))

        # Main scrollable/padded container
        container = tk.Frame(self.root, bg="#0F172A", padx=20, pady=15)
        container.pack(fill="both", expand=True)

        # 2. Main File Selection Card
        main_card = tk.LabelFrame(
            container, 
            text=" 1. Main Target Files (Excel / CSV / DBF) ", 
            font=("Segoe UI", 10, "bold"), 
            fg="#F8FAFC", 
            bg="#1E293B",
            padx=15, 
            pady=12,
            bd=1,
            relief="solid"
        )
        main_card.pack(fill="x", pady=(0, 15))

        main_desc = tk.Label(
            main_card, 
            text="Select the main spreadsheet(s) or database table(s) where ABC_IDs need to be matched.",
            font=("Segoe UI", 9),
            fg="#94A3B8",
            bg="#1E293B"
        )
        main_desc.pack(anchor="w", pady=(0, 8))

        main_action_frame = tk.Frame(main_card, bg="#1E293B")
        main_action_frame.pack(fill="x", pady=(0, 8))

        self.btn_select_main = self.make_flat_button(
            main_action_frame, 
            "Add Main File(s)", 
            self.browse_main_file,
            bg="#3B82F6",
            hover_bg="#2563EB"
        )
        self.btn_select_main.pack(side="left")

        self.btn_clear_main = self.make_flat_button(
            main_action_frame, 
            "Clear List", 
            self.clear_main_file,
            bg="#64748B",
            hover_bg="#475569"
        )
        self.btn_clear_main.pack(side="left", padx=(8, 0))

        self.lbl_main_file = tk.Label(
            main_action_frame,
            text="No main file selected.",
            font=("Segoe UI", 9, "italic"),
            fg="#EF4444",
            bg="#1E293B",
            padx=10
        )
        self.lbl_main_file.pack(side="left", fill="x", expand=True, anchor="w")

        # Scrolled text list of main files
        self.main_files_box = ScrolledText(
            main_card,
            height=3,
            bg="#0F172A",
            fg="#E2E8F0",
            insertbackground="#E2E8F0",
            font=("Consolas", 9),
            bd=1,
            relief="solid"
        )
        self.main_files_box.pack(fill="x", expand=True)
        self.main_files_box.insert("end", "No main files selected.\n")
        self.main_files_box.config(state="disabled")

        # 3. Master Files Selection Card
        master_card = tk.LabelFrame(
            container, 
            text=" 2. Master ABC ID Data Files (Multiple Allowed) ", 
            font=("Segoe UI", 10, "bold"), 
            fg="#F8FAFC", 
            bg="#1E293B",
            padx=15, 
            pady=12,
            bd=1,
            relief="solid"
        )
        master_card.pack(fill="x", pady=(0, 15))

        master_desc = tk.Label(
            master_card, 
            text="Select one or multiple Excel, CSV, or DBF files containing the source PRN -> ABC_ID mapping dictionary.",
            font=("Segoe UI", 9),
            fg="#94A3B8",
            bg="#1E293B"
        )
        master_desc.pack(anchor="w", pady=(0, 8))

        master_action_frame = tk.Frame(master_card, bg="#1E293B")
        master_action_frame.pack(fill="x", pady=(0, 8))

        self.btn_select_master = self.make_flat_button(
            master_action_frame, 
            "Add Master File(s)", 
            self.browse_master_files,
            bg="#10B981",
            hover_bg="#059669"
        )
        self.btn_select_master.pack(side="left")

        self.btn_clear_master = self.make_flat_button(
            master_action_frame, 
            "Clear List", 
            self.clear_master_files,
            bg="#64748B",
            hover_bg="#475569"
        )
        self.btn_clear_master.pack(side="left", padx=(8, 0))

        # Scrolled text list of master files
        self.master_files_box = ScrolledText(
            master_card,
            height=3,
            bg="#0F172A",
            fg="#E2E8F0",
            insertbackground="#E2E8F0",
            font=("Consolas", 9),
            bd=1,
            relief="solid"
        )
        self.master_files_box.pack(fill="x", expand=True)
        self.master_files_box.insert("end", "No master files added yet.\n")
        self.master_files_box.config(state="disabled")

        # 4. Output Configuration Card
        output_card = tk.LabelFrame(
            container, 
            text=" 3. Output Configuration & Options ", 
            font=("Segoe UI", 10, "bold"), 
            fg="#F8FAFC", 
            bg="#1E293B",
            padx=15, 
            pady=12,
            bd=1,
            relief="solid"
        )
        output_card.pack(fill="x", pady=(0, 15))

        output_action_frame = tk.Frame(output_card, bg="#1E293B")
        output_action_frame.pack(fill="x", pady=(0, 8))

        self.btn_select_output = self.make_flat_button(
            output_action_frame, 
            "Select Output Folder", 
            self.browse_output_dir,
            bg="#F59E0B",
            hover_bg="#D97706"
        )
        self.btn_select_output.pack(side="left")

        self.lbl_output_dir = tk.Label(
            output_action_frame,
            text="No output folder selected...",
            font=("Segoe UI", 9, "italic"),
            fg="#EF4444",
            bg="#1E293B",
            padx=10
        )
        self.lbl_output_dir.pack(side="left", fill="x", expand=True, anchor="w")

        options_frame = tk.Frame(output_card, bg="#1E293B")
        options_frame.pack(fill="x")

        lbl_mode = tk.Label(
            options_frame,
            text="Output Mode:",
            font=("Segoe UI", 9, "bold"),
            fg="#F8FAFC",
            bg="#1E293B",
            pady=4
        )
        lbl_mode.pack(side="left", anchor="w")

        self.radio_sep = tk.Radiobutton(
            options_frame,
            text="Separate by Degree Bucket",
            variable=self.output_mode,
            value="separate_degree",
            bg="#1E293B",
            fg="#F8FAFC",
            selectcolor="#0F172A",
            activebackground="#1E293B",
            activeforeground="#F8FAFC",
            font=("Segoe UI", 9),
            padx=10
        )
        self.radio_sep.pack(side="left")

        self.radio_merge = tk.Radiobutton(
            options_frame,
            text="Consolidate into Single File",
            variable=self.output_mode,
            value="merge_single",
            bg="#1E293B",
            fg="#F8FAFC",
            selectcolor="#0F172A",
            activebackground="#1E293B",
            activeforeground="#F8FAFC",
            font=("Segoe UI", 9),
            padx=10
        )
        self.radio_merge.pack(side="left")

        self.radio_indiv = tk.Radiobutton(
            options_frame,
            text="Process Individually",
            variable=self.output_mode,
            value="individual",
            bg="#1E293B",
            fg="#F8FAFC",
            selectcolor="#0F172A",
            activebackground="#1E293B",
            activeforeground="#F8FAFC",
            font=("Segoe UI", 9),
            padx=10
        )
        self.radio_indiv.pack(side="left")

        # 5. Action & Console Card
        console_card = tk.LabelFrame(
            container, 
            text=" 4. Matching Dashboard & Logs ", 
            font=("Segoe UI", 10, "bold"), 
            fg="#F8FAFC", 
            bg="#1E293B",
            padx=15, 
            pady=12,
            bd=1,
            relief="solid"
        )
        console_card.pack(fill="both", expand=True)

        action_frame = tk.Frame(console_card, bg="#1E293B")
        action_frame.pack(fill="x", pady=(0, 10))

        self.btn_run = self.make_flat_button(
            action_frame, 
            "Start Matching & Update", 
            self.start_processing_thread,
            bg="#6366F1",
            hover_bg="#4F46E5",
            font=("Segoe UI", 10, "bold")
        )
        self.btn_run.pack(side="left")
        self.btn_run.config(state="disabled")

        self.lbl_status = tk.Label(
            action_frame,
            text="Status: Waiting for files...",
            font=("Segoe UI", 9, "bold"),
            fg="#94A3B8",
            bg="#1E293B",
            padx=15
        )
        self.lbl_status.pack(side="left")

        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            console_card, 
            variable=self.progress_var, 
            maximum=100,
            style="Indigo.Horizontal.TProgressbar"
        )
        self.progress_bar.pack(fill="x", pady=(0, 10))

        # Log Console
        self.console_box = ScrolledText(
            console_card,
            height=6,
            bg="#020617",
            fg="#38BDF8",
            insertbackground="#38BDF8",
            font=("Consolas", 9),
            bd=0
        )
        self.console_box.pack(fill="both", expand=True)
        self.log("Console initialized. Ready for operations.", "SYSTEM")

    # Log Helper
    def log(self, text, tag="INFO"):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.msg_queue.put(("log", f"[{timestamp}] [{tag}] {text}\n"))

    def set_progress(self, val):
        self.msg_queue.put(("progress", val))

    def set_status(self, text, color="#94A3B8"):
        self.msg_queue.put(("status", (text, color)))

    # Browse Button Callbacks
    def browse_main_file(self):
        file_types = [
            ("All Supported Formats", "*.xlsx;*.csv;*.dbf"),
            ("Excel Spreadsheets (*.xlsx)", "*.xlsx"),
            ("CSV Text Files (*.csv)", "*.csv"),
            ("dBase Database Files (*.dbf)", "*.dbf")
        ]
        path = filedialog.askopenfilename(title="Select Main Target File", filetypes=file_types)
        if path:
            self.main_file_path = os.path.abspath(path)
            self.lbl_main_file.config(text=self.main_file_path, fg="#34D399") # Green success
            self.log(f"Main target file selected: {self.main_file_path}")
            self.check_ready_state()

    def browse_master_files(self):
        file_types = [
            ("All Supported Formats", "*.xlsx;*.csv;*.dbf;*.xls"),
            ("Excel Spreadsheets (*.xlsx;*.xls)", "*.xlsx;*.xls"),
            ("CSV Text Files (*.csv)", "*.csv"),
            ("dBase Database Files (*.dbf)", "*.dbf")
        ]
        paths = filedialog.askopenfilenames(title="Select Master Data Files (Select Multiple)", filetypes=file_types)
        if paths:
            for p in paths:
                abs_p = os.path.abspath(p)
                if abs_p not in self.master_files_paths:
                    self.master_files_paths.append(abs_p)
                    self.log(f"Added master file source: {abs_p}")
            self.update_master_files_box()
            self.check_ready_state()

    def clear_master_files(self):
        self.master_files_paths = []
        self.update_master_files_box()
        self.log("Master files list cleared.")
        self.check_ready_state()

    def clear_main_file(self):
        self.main_file_path = ""
        self.lbl_main_file.config(text="No main file selected.", fg="#94A3B8")
        self.log("Main file cleared.")
        self.check_ready_state()

    def browse_output_dir(self):
        directory = filedialog.askdirectory(title="Select Output Directory")
        if directory:
            self.output_dir_path = os.path.abspath(directory)
            self.lbl_output_dir.config(text=self.output_dir_path, fg="#34D399")
            self.log(f"Output directory selected: {self.output_dir_path}")
            self.check_ready_state()

    def update_master_files_box(self):
        self.master_files_box.config(state="normal")
        self.master_files_box.delete("1.0", "end")
        if not self.master_files_paths:
            self.master_files_box.insert("end", "No master files added yet.\n")
        else:
            for i, p in enumerate(self.master_files_paths, 1):
                self.master_files_box.insert("end", f"[{i}] {p}\n")
        self.master_files_box.config(state="disabled")

    def check_ready_state(self):
        """Enable or disable the execute button depending on selection states."""
        if self.main_file_path and self.master_files_paths:
            self.btn_run.config(state="normal", bg="#6366F1")
            self.set_status("Ready to run matching process.", "#34D399")
        else:
            self.btn_run.config(state="disabled", bg="#475569")
            if not self.main_file_path:
                self.set_status("Waiting for main file selection...", "#94A3B8")
            elif not self.master_files_paths:
                self.set_status("Waiting for master files selection...", "#94A3B8")

    # Queue Poller for Thread Safety
    def process_queue(self):
        try:
            while True:
                msg_type, data = self.msg_queue.get_nowait()
                if msg_type == "log":
                    self.console_box.insert("end", data)
                    self.console_box.see("end")
                elif msg_type == "progress":
                    self.progress_var.set(data)
                elif msg_type == "status":
                    text, color = data
                    self.lbl_status.config(text=text, fg=color)
                elif msg_type == "finished":
                    self.enable_controls(True)
                    self.show_completion_summary(data)
                elif msg_type == "error":
                    self.enable_controls(True)
                    messagebox.showerror("Error Occurred", data)
                self.msg_queue.task_done()
        except queue.Empty:
            pass
        self.root.after(100, self.process_queue)

    def enable_controls(self, state):
        flag = "normal" if state else "disabled"
        self.btn_select_main.config(state=flag)
        self.btn_select_master.config(state=flag)
        self.btn_clear_master.config(state=flag)
        self.btn_run.config(state=flag)

    def start_processing_thread(self):
        self.enable_controls(False)
        self.progress_var.set(0)
        self.set_status("Processing...", "#F59E0B")
        
        # Start matching thread
        threading.Thread(target=self.run_matching, daemon=True).start()

    def run_matching(self):
        try:
            self.log("Starting ABC ID Matcher thread...", "PROCESS")

            # 1. Compile PRN -> ABC_ID dictionary from master files
            prn_to_abc = {}
            total_master_files = len(self.master_files_paths)
            for index, m_path in enumerate(self.master_files_paths):
                m_base = os.path.basename(m_path)
                self.log(f"Reading master file ({index+1}/{total_master_files}): {m_base}...", "MASTER")
                self.set_status(f"Loading master file {index+1}/{total_master_files}...", "#F59E0B")
                
                # Load master data depending on type
                ext = os.path.splitext(m_path.lower())[1]
                records_loaded = 0
                
                try:
                    if ext == '.dbf':
                        db = DBF(m_path, ignore_missing_memofile=True)
                        fields = db.field_names
                        prn_col, abc_col = find_matching_columns(fields)
                        if not prn_col or not abc_col:
                            self.log(f"Could not map columns in DBF file: {m_base}. Found fields: {fields}", "WARNING")
                            continue
                        for record in db:
                            prn_val = clean_value(record.get(prn_col))
                            abc_val = clean_value(record.get(abc_col))
                            if prn_val and abc_val:
                                prn_to_abc[prn_val] = abc_val
                                records_loaded += 1
                                
                    elif ext in ['.xlsx', '.xls']:
                        with pd.ExcelFile(m_path) as xls:
                            for sheet in xls.sheet_names:
                                df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
                                prn_col, abc_col = find_matching_columns(df.columns)
                                if not prn_col or not abc_col:
                                    self.log(f"Skip sheet '{sheet}' in Excel: Column mapping failed. Found: {list(df.columns)}", "WARNING")
                                    continue
                                for _, row in df.iterrows():
                                    prn_val = clean_value(row[prn_col])
                                    abc_val = clean_value(row[abc_col])
                                    if prn_val and abc_val:
                                        prn_to_abc[prn_val] = abc_val
                                        records_loaded += 1
                                    
                    elif ext == '.csv':
                        try:
                            df = pd.read_csv(m_path, dtype=str, encoding='utf-8-sig', on_bad_lines='skip')
                        except Exception:
                            df = pd.read_csv(m_path, dtype=str, encoding='cp1252', on_bad_lines='skip')
                        prn_col, abc_col = find_matching_columns(df.columns)
                        if not prn_col or not abc_col:
                            self.log(f"Could not map columns in CSV: {m_base}. Found: {list(df.columns)}", "WARNING")
                            continue
                        for _, row in df.iterrows():
                            prn_val = clean_value(row[prn_col])
                            abc_val = clean_value(row[abc_col])
                            if prn_val and abc_val:
                                prn_to_abc[prn_val] = abc_val
                                records_loaded += 1
                    else:
                        self.log(f"Unsupported file format for master data: {ext}", "ERROR")
                        continue
                        
                    self.log(f"Loaded {records_loaded} mapping records from {m_base}.", "MASTER")
                except Exception as ex:
                    self.log(f"Failed to read master file {m_base}: {str(ex)}", "ERROR")
                    continue
                    
                progress = int(((index + 1) / total_master_files) * 30)
                self.set_progress(progress)

            self.log(f"Master files loading complete. Total unique mappings compiled: {len(prn_to_abc)}", "SUCCESS")
            if not prn_to_abc:
                raise ValueError("No matching PRN to ABC ID mappings could be parsed from master files.")

            # 2. Determine processing mode
            mode = self.output_mode.get()
            self.log(f"Processing mode: {mode}", "PROCESS")
            
            stats = {
                "total_rows": 0,
                "updated_rows": 0,
                "unmatched_rows": 0,
                "output_dir": self.output_dir_path,
                "mode": mode,
                "files_count": len(self.main_files_paths),
                "generated_files": []
            }

            if mode == "individual":
                # Process each file and save updated copy to output directory
                for idx, main_file in enumerate(self.main_files_paths):
                    main_base = os.path.basename(main_file)
                    self.log(f"Processing main file ({idx+1}/{len(self.main_files_paths)}): {main_base}", "MAIN")
                    self.set_status(f"Processing file {idx+1}/{len(self.main_files_paths)}...", "#F59E0B")
                    
                    out_path = os.path.join(self.output_dir_path, main_base)
                    
                    # Copy to output directory first
                    shutil.copy2(main_file, out_path)
                    
                    main_ext = os.path.splitext(main_base.lower())[1]
                    if main_ext == '.xlsx':
                        wb = openpyxl.load_workbook(out_path, data_only=False)
                        ws = wb.active
                        header_row_idx = 1
                        prn_col_idx = None
                        abc_col_idx = None
                        
                        for r in range(1, min(15, ws.max_row + 1)):
                            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                            headers_temp = [str(v).strip() if v is not None else "" for v in row_vals]
                            p_col, a_col = find_matching_columns(headers_temp)
                            if p_col is not None:
                                header_row_idx = r
                                for c in range(1, ws.max_column + 1):
                                    val = str(ws.cell(row=r, column=c).value).strip() if ws.cell(row=r, column=c).value is not None else ""
                                    if val == p_col:
                                        prn_col_idx = c
                                    if a_col is not None and val == a_col:
                                        abc_col_idx = c
                                break

                        if prn_col_idx is None:
                            wb.close()
                            raise ValueError(f"Could not find a column mapping for 'PRN' in sheet '{ws.title}' first 15 rows.")

                        if abc_col_idx is None:
                            abc_col_idx = ws.max_column + 1
                            ws.cell(row=header_row_idx, column=abc_col_idx, value="ABC_ID")
                            prn_hdr_cell = ws.cell(row=header_row_idx, column=prn_col_idx)
                            abc_hdr_cell = ws.cell(row=header_row_idx, column=abc_col_idx)
                            if prn_hdr_cell.has_style:
                                from copy import copy
                                abc_hdr_cell.font = copy(prn_hdr_cell.font)
                                abc_hdr_cell.fill = copy(prn_hdr_cell.fill)
                                abc_hdr_cell.border = copy(prn_hdr_cell.border)
                                abc_hdr_cell.alignment = copy(prn_hdr_cell.alignment)

                        for r in range(header_row_idx + 1, ws.max_row + 1):
                            prn_val = ws.cell(row=r, column=prn_col_idx).value
                            if prn_val is None or str(prn_val).strip() == "":
                                continue
                            existing_abc = ws.cell(row=r, column=abc_col_idx).value
                            if existing_abc is not None and str(existing_abc).strip() != "":
                                continue
                            
                            stats["total_rows"] += 1
                            cleaned_prn = clean_value(prn_val)
                            if cleaned_prn in prn_to_abc:
                                target_cell = ws.cell(row=r, column=abc_col_idx)
                                target_cell.value = str(prn_to_abc[cleaned_prn])
                                target_cell.number_format = '@'
                                stats["updated_rows"] += 1
                            else:
                                stats["unmatched_rows"] += 1
                                
                        wb.save(out_path)
                        wb.close()
                        
                    elif main_ext == '.csv':
                        encoding = 'utf-8-sig'
                        delimiter = ','
                        try:
                            with open(out_path, 'r', encoding='utf-8-sig') as f:
                                sample = f.read(4096)
                                f.seek(0)
                                if sample:
                                    try:
                                        dialect = csv.Sniffer().sniff(sample)
                                        delimiter = dialect.delimiter
                                    except Exception:
                                        pass
                        except Exception:
                            encoding = 'cp1252'

                        rows = []
                        with open(out_path, 'r', encoding=encoding, newline='') as f:
                            reader = csv.reader(f, delimiter=delimiter)
                            for r in reader:
                                rows.append(r)

                        if not rows:
                            raise ValueError("CSV target file is empty.")

                        header_row_idx = 0
                        prn_col_idx = None
                        abc_col_idx = None
                        
                        for r_idx in range(min(15, len(rows))):
                            headers_temp = [str(h).strip() for h in rows[r_idx]]
                            p_col, a_col = find_matching_columns(headers_temp)
                            if p_col is not None:
                                header_row_idx = r_idx
                                for c, h in enumerate(rows[r_idx]):
                                    h_clean = h.strip()
                                    if h_clean == p_col:
                                        prn_col_idx = c
                                    if a_col is not None and h_clean == a_col:
                                        abc_col_idx = c
                                break

                        if prn_col_idx is None:
                            raise ValueError("Could not find a 'PRN' column in the CSV file header row.")

                        if abc_col_idx is None:
                            abc_col_idx = len(rows[header_row_idx])
                            rows[header_row_idx].append("ABC_ID")
                            for r_idx in range(header_row_idx):
                                rows[r_idx].append("")

                        for r_idx in range(header_row_idx + 1, len(rows)):
                            row = rows[r_idx]
                            while len(row) < abc_col_idx + 1:
                                row.append("")
                            prn_val = row[prn_col_idx]
                            if not prn_val or str(prn_val).strip() == "":
                                continue
                            existing_abc = row[abc_col_idx]
                            if existing_abc is not None and str(existing_abc).strip() != "":
                                continue
                            
                            stats["total_rows"] += 1
                            cleaned_prn = clean_value(prn_val)
                            if cleaned_prn in prn_to_abc:
                                row[abc_col_idx] = prn_to_abc[cleaned_prn]
                                stats["updated_rows"] += 1
                            else:
                                stats["unmatched_rows"] += 1

                        with open(out_path, 'w', encoding=encoding, newline='') as f:
                            writer = csv.writer(f, delimiter=delimiter)
                            writer.writerows(rows)

                    elif main_ext == '.dbf':
                        table = dbf.Table(out_path)
                        table.open(mode=dbf.READ_WRITE)
                        fields = table.field_names
                        prn_col, abc_col = find_matching_columns(fields)
                        if prn_col is None:
                            table.close()
                            raise ValueError(f"Could not find a PRN column in DBF table fields: {fields}")

                        if abc_col is None:
                            self.log("Field 'ABCID' is missing in target DBF structure. Adding it C(20)...", "DBF")
                            table.add_fields('ABCID C(20)')
                            table.open(mode=dbf.READ_WRITE)
                            abc_col = 'ABCID'

                        for record in table:
                            prn_val = record[prn_col]
                            if prn_val is None or str(prn_val).strip() == "":
                                continue
                            existing_abc = record[abc_col]
                            if existing_abc is not None and str(existing_abc).strip() != "":
                                continue
                            
                            stats["total_rows"] += 1
                            cleaned_prn = clean_value(prn_val)
                            if cleaned_prn in prn_to_abc:
                                dbf.write(record, **{abc_col: prn_to_abc[cleaned_prn]})
                                stats["updated_rows"] += 1
                            else:
                                stats["unmatched_rows"] += 1
                        table.close()

                    self.log(f"Successfully processed and saved copy to output: {main_base}", "MAIN")
                    stats["generated_files"].append(out_path)
                    
                    prog = 30 + int(((idx + 1) / len(self.main_files_paths)) * 70)
                    self.set_progress(prog)

            else:
                # Merge or Separate by Degree Bucket
                # 2.1 Load Program Master
                script_dir = Path(__file__).parent.resolve()
                bucket_master_path = script_dir / DEFAULT_BUCKET_MASTER
                program_master_path = script_dir / "program_master.xlsx"
                master = load_bucket_master(bucket_master_path)
                if program_master_path.exists():
                    self.log(f"Syncing master mappings from {program_master_path.name}...")
                    sync_master_from_program_master(master, program_master_path)
                    save_bucket_master(bucket_master_path, master)
                else:
                    rebuild_master_indexes(master)

                # 2.2 Load all rows from all main files
                consolidated_rows = []
                union_headers = []
                
                for idx, main_file in enumerate(self.main_files_paths):
                    main_base = os.path.basename(main_file)
                    self.log(f"Reading rows from: {main_base}", "MAIN")
                    file_headers, file_rows = read_rows_from_file(main_file)
                    
                    # Update union headers
                    for h in file_headers:
                        if h not in union_headers:
                            union_headers.append(h)
                            
                    # Find PRN and ABC columns in this file
                    prn_col, abc_col = find_matching_columns(file_headers)
                    if not prn_col:
                        raise ValueError(f"Could not find PRN column in file: {main_base}")
                        
                    # If ABC ID column is not in headers, add it
                    if not abc_col:
                        abc_col = "ABC_ID"
                        if abc_col not in union_headers:
                            union_headers.append(abc_col)

                    # Store info on how to process
                    for row in file_rows:
                        # Normalize keys to headers
                        row_normalized = {k: row.get(k, "") for k in file_headers}
                        if abc_col not in row_normalized:
                            row_normalized[abc_col] = ""
                            
                        # Add ABC_ID match in-place
                        stats["total_rows"] += 1
                        prn_val = row_normalized.get(prn_col)
                        cleaned_prn = clean_value(prn_val)
                        existing_abc = row_normalized.get(abc_col)
                        
                        if prn_val and (existing_abc is None or str(existing_abc).strip() == ""):
                            if cleaned_prn in prn_to_abc:
                                row_normalized[abc_col] = prn_to_abc[cleaned_prn]
                                stats["updated_rows"] += 1
                            else:
                                stats["unmatched_rows"] += 1
                        else:
                            if existing_abc:
                                # Count as already matched or processed, but not updated now
                                pass
                                
                        row_normalized["_source_file"] = main_base
                        consolidated_rows.append((row_normalized, prn_col, abc_col))

                    prog = 30 + int(((idx + 1) / len(self.main_files_paths)) * 30)
                    self.set_progress(prog)

                if "ABC_ID" not in union_headers:
                    union_headers.append("ABC_ID")

                # Remove any internal fields from output column order
                if "_source_file" in union_headers:
                    union_headers.remove("_source_file")

                if mode == "merge_single":
                    # Output a single consolidated workbook
                    self.set_status("Saving merged workbook...", "#F59E0B")
                    out_path = os.path.join(self.output_dir_path, "consolidated_matched.xlsx")
                    
                    # Extract pure dicts
                    pure_rows = [r[0] for r in consolidated_rows]
                    
                    self.log("Writing consolidated rows to Excel...", "SAVE")
                    write_bucket_workbook(Path(out_path), DEFAULT_SHEET_NAME, union_headers, pure_rows)
                    stats["generated_files"].append(out_path)
                    
                elif mode == "separate_degree":
                    # Separate by Degree Bucket
                    self.set_status("Separating by degree buckets...", "#F59E0B")
                    rows_by_bucket = defaultdict(list)
                    
                    for row_tuple in consolidated_rows:
                        row_dict, prn_col, abc_col = row_tuple
                        
                        # Find values for mapping resolution
                        appl_no = row_dict.get("APPL_NO") or row_dict.get("PROG_NO") or row_dict.get("appl_no") or row_dict.get("prog_no")
                        abbr = row_dict.get("ABBR") or row_dict.get("abbr")
                        degnm = row_dict.get("DEGNM") or row_dict.get("degnm")
                        
                        bucket_name, _ = resolve_row_from_indexes(
                            master, program_no=appl_no, abbr=abbr, degree_name=degnm
                        )
                        
                        if not bucket_name:
                            # Pause matching thread and request manual mapping from main thread
                            mapping_event = threading.Event()
                            mapping_result = []
                            
                            self.msg_queue.put(("request_mapping", {
                                "row_dict": row_dict,
                                "file_name": row_dict.get("_source_file", "Main File"),
                                "event": mapping_event,
                                "result": mapping_result
                            }))
                            
                            # Wait for UI input
                            mapping_event.wait()
                            
                            if not mapping_result:
                                raise ValueError("Matching process cancelled during manual bucket selection.")
                            bucket_name = mapping_result[0]
                            
                            # Reload master since it was updated and saved by the dialog thread
                            master = load_bucket_master(bucket_master_path)
                            
                        rows_by_bucket[bucket_name].append(row_dict)

                    # Save all bucketted spreadsheets
                    self.log("Writing degree bucket files to Excel...", "SAVE")
                    for bucket_name, rows in rows_by_bucket.items():
                        if not rows: continue
                        out_name = f"{safe_filename(bucket_name)}.xlsx"
                        out_path = os.path.join(self.output_dir_path, out_name)
                        write_bucket_workbook(Path(out_path), DEFAULT_SHEET_NAME, union_headers, rows)
                        stats["generated_files"].append(out_path)

            self.set_progress(100)
            self.set_status("Execution Completed Successfully!", "#34D399")
            self.msg_queue.put(("finished", stats))

        except Exception as err:
            self.log(f"ERROR: {str(err)}", "ERROR")
            self.set_status("Error Occurred", "#EF4444")
            self.msg_queue.put(("error", str(err)))

    # Summary completion dialog
    def show_completion_summary(self, stats):
        # Create a beautiful custom summary window
        summary_win = tk.Toplevel(self.root)
        summary_win.title("Matching Run Complete")
        summary_win.geometry("540x480")
        summary_win.resizable(False, False)
        summary_win.configure(bg="#0F172A")
        summary_win.transient(self.root)
        summary_win.grab_set()

        # Layout inside summary
        padded_frame = tk.Frame(summary_win, bg="#0F172A", padx=25, pady=20)
        padded_frame.pack(fill="both", expand=True)

        header_lbl = tk.Label(
            padded_frame,
            text="Matching Complete!",
            font=("Segoe UI", 16, "bold"),
            fg="#10B981", # Emerald
            bg="#0F172A"
        )
        header_lbl.pack(pady=(0, 15), anchor="w")

        info_lbl = tk.Label(
            padded_frame,
            text="The matching run finished successfully. Detailed stats are listed below:",
            font=("Segoe UI", 10),
            fg="#94A3B8",
            bg="#0F172A",
            wraplength=490,
            justify="left"
        )
        info_lbl.pack(pady=(0, 15), anchor="w")

        # Stats Card
        stats_card = tk.Frame(padded_frame, bg="#1E293B", padx=15, pady=15, bd=1, relief="solid")
        stats_card.pack(fill="x", pady=(0, 20))

        # Helper row creator
        def add_stat_row(parent, label, value, value_color="#F8FAFC"):
            row = tk.Frame(parent, bg="#1E293B")
            row.pack(fill="x", pady=4)
            lbl = tk.Label(row, text=label, font=("Segoe UI", 9, "bold"), fg="#94A3B8", bg="#1E293B")
            lbl.pack(side="left")
            val = tk.Label(row, text=value, font=("Segoe UI", 10, "bold"), fg=value_color, bg="#1E293B")
            val.pack(side="right")

        total = stats["total_rows"]
        updated = stats["updated_rows"]
        unmatched = stats["unmatched_rows"]
        
        pct_matched = f"{(updated / total * 100):.1f}%" if total > 0 else "0.0%"

        mode_names = {
            "separate_degree": "Separate by Degree Bucket",
            "merge_single": "Consolidate into Single File",
            "individual": "Process Individually"
        }

        add_stat_row(stats_card, "Processing Mode:", mode_names.get(stats["mode"], stats["mode"]), "#818CF8")
        add_stat_row(stats_card, "Input Files Loaded:", f"{stats['files_count']} file(s)")
        add_stat_row(stats_card, "Total Rows Evaluated:", str(total))
        add_stat_row(stats_card, "ABC ID Found & Updated:", f"{updated} ({pct_matched})", "#34D399")
        add_stat_row(stats_card, "Missing in Master Files:", str(unmatched), "#EF4444")
        add_stat_row(stats_card, "Output Files Generated:", f"{len(stats['generated_files'])} file(s)", "#F59E0B")

        # Action Buttons
        btn_frame = tk.Frame(padded_frame, bg="#0F172A")
        btn_frame.pack(fill="x", pady=(5, 0))

        def open_folder():
            # Open output folder location in Windows Explorer
            os.startfile(stats["output_dir"])

        btn_folder = self.make_flat_button(
            btn_frame,
            "Open Output Directory",
            open_folder,
            bg="#3B82F6",
            hover_bg="#2563EB"
        )
        btn_folder.pack(side="left")

        btn_close = self.make_flat_button(
            btn_frame,
            "OK, Close",
            summary_win.destroy,
            bg="#6366F1",
            hover_bg="#4F46E5"
        )
        btn_close.pack(side="right")


if __name__ == "__main__":
    # Ensure Tkinter runs correctly
    root = tk.Tk()
    app = ABCIDMatcherApp(root)
    root.mainloop()
