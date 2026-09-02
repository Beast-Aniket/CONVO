#!/usr/bin/env python3
"""Merge Excel files into bucket workbooks using an editable Excel master."""

from __future__ import annotations

import argparse
import re
from collections import OrderedDict, defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

STANDARD_COLUMNS = [
    "LotNo",
    "Conv ID",
    "Faculty",
    "PRNERN",
    "ProgType",
    "APPL_NO",
    "SEAT_NO",
    "COLL_NO",
    "COLL_NAME",
    "COLL_NAMEM",
    "StudLastName",
    "StudFirstName",
    "StudMidddleName",
    "StudMotherName",
    "NAME",
    "NAME_MARAT",
    "SEX",
    "ABBR",
    "CLASS",
    "MCLASS",
    "SUB1",
    "SUB1_NAME",
    "SUB1_NAMEM",
    "SUB2",
    "SUB2_NAME",
    "SUB2_NAMEM",
    "DEGNM",
    "MDEGNM",
    "SUBDEGNM",
    "MSUBDEGNM",
    "PER",
    "MPER",
]

TEXT_ONLY_COLUMNS = {"PRNERN", "MCLASS"}
CLASS_COLUMN = "CLASS"
DEFAULT_SHEET_NAME = "Structured"
DEFAULT_BUCKET_MASTER = "program_bucket_master.xlsx"
SHEET_NOTES = "Read_Me"
SHEET_BUCKETS = "Bucket_List"
SHEET_DEGREES = "Degree_List"
SHEET_PROGRAMS = "Program_Mapping"

DEFAULT_BUCKETS = [
    ("COMMERCE", "convocation_commerce@mu.ac.in"),
    ("MANAGEMENT", "convocation_management@mu.ac.in"),
    ("SCIENCE", "convocation_Science@mu.ac.in"),
    ("ENGINEERING", "convocation_engineering@mu.ac.in"),
    ("ART", "convocation_art@mu.ac.in"),
    ("LAW", "convocation_law@mu.ac.in"),
]

PROGRAM_MASTER_COLUMNS = [
    "PROG_NO",
    "ABBR",
    "FACULTY",
    "DEGNM",
    "MDEGNM",
    "SUBDEGNM",
    "MSUBDEGNM",
    "MERGE_NAME",
    "EMAIL",
    "SOURCE",
    "STATUS",
]

DEGREE_MASTER_COLUMNS = [
    "DEGNM",
    "MERGE_NAME",
    "EMAIL",
    "SOURCE",
    "STATUS",
    "PROGRAM_COUNT",
    "FACULTIES",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Merge Excel files into editable bucket-based outputs using "
            "program_bucket_master.xlsx as the mapping source."
        )
    )
    parser.add_argument(
        "--input-dir",
        help="Folder containing source .xlsx files for no-GUI mode.",
    )
    parser.add_argument(
        "--input-files",
        nargs="*",
        help="Specific source .xlsx files for no-GUI mode.",
    )
    parser.add_argument(
        "--output-dir",
        help="Folder where merged files will be written.",
    )
    parser.add_argument(
        "--program-master",
        help="Path to the raw program master workbook.",
    )
    parser.add_argument(
        "--bucket-master",
        help="Path to the editable Excel bucket master workbook.",
    )
    parser.add_argument(
        "--sheet-name",
        default=DEFAULT_SHEET_NAME,
        help=f"Worksheet to read from each source file. Default: {DEFAULT_SHEET_NAME}",
    )
    parser.add_argument(
        "--build-master-only",
        action="store_true",
        help="Create or refresh the editable bucket master workbook and exit.",
    )
    parser.add_argument(
        "--no-gui",
        action="store_true",
        help="Run without file dialogs. Useful for testing or batch runs.",
    )
    return parser.parse_args()


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_lookup_key(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def is_blank_row(values: tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def as_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        text = format(value, ".15g")
        if "e" in text.lower():
            return format(Decimal(str(value)), "f").rstrip("0").rstrip(".")
        return text
    return str(value)


def parse_numeric_class(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return float(Decimal(text))
        except InvalidOperation:
            return None
    return None


def safe_filename(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\\\|?*]+', "_", name.strip())
    cleaned = re.sub(r"\s+", "_", cleaned)
    cleaned = re.sub(r"_+", "_", cleaned).strip("._")
    return cleaned or "OUTPUT"


def get_default_sheet_name(workbook, preferred_sheet_name: str) -> str:
    if preferred_sheet_name in workbook.sheetnames:
        return preferred_sheet_name
    for name in workbook.sheetnames:
        if normalize_lookup_key(name) != "AUDIT_LOG":
            return name
    return workbook.sheetnames[0]


def classify_degree_name(degree_name: Any) -> tuple[str | None, str]:
    degree_key = normalize_lookup_key(degree_name)
    if not degree_key:
        return None, "EMPTY_DEGREE"

    if re.search(r"\bLAWS?\b", degree_key):
        return "LAW", "AUTO:LAW"
    if "BUSINESS ADMINISTRATION" in degree_key or re.search(
        r"\bMANAGEMENT\b", degree_key
    ):
        return "MANAGEMENT", "AUTO:MANAGEMENT"
    if re.search(r"\bCOMMERCE\b", degree_key):
        return "COMMERCE", "AUTO:COMMERCE"
    if re.search(r"\bARCHITECTURE\b", degree_key):
        return "SCIENCE", "AUTO:ARCHITECTURE_TO_SCIENCE"
    if re.search(r"\bENGINEERING\b", degree_key):
        return "ENGINEERING", "AUTO:ENGINEERING"
    if "COMPUTER APPLICATIONS" in degree_key or re.search(
        r"\bSCIENCE\b|\bPHARMACY\b", degree_key
    ):
        return "SCIENCE", "AUTO:SCIENCE"
    if (
        "FINE ARTS" in degree_key
        or "PHYSICAL EDUCATION" in degree_key
        or "SOCIAL WORK" in degree_key
        or "EDUCATION" in degree_key
        or "MUSIC" in degree_key
        or re.search(r"\bARTS?\b", degree_key)
    ):
        return "ART", "AUTO:ART"
    return None, "REVIEW_NEEDED"


def create_empty_master() -> dict[str, Any]:
    master = {
        "buckets": OrderedDict(),
        "degree_rows": OrderedDict(),
        "program_rows": OrderedDict(),
        "program_master_path": "",
        "indexes": {
            "degrees": {},
            "programs_by_prog": {},
            "programs_by_abbr": {},
        },
    }
    seed_default_buckets(master)
    return master


def seed_default_buckets(master: dict[str, Any]) -> None:
    for bucket_name, email in DEFAULT_BUCKETS:
        add_bucket(master, bucket_name, email)


def add_bucket(master: dict[str, Any], bucket_name: Any, email: Any = "") -> str:
    display_name = str(bucket_name).strip()
    if not display_name:
        return ""
    key = normalize_lookup_key(display_name)
    existing = master["buckets"].get(key)
    if existing:
        if str(email).strip():
            existing["email"] = str(email).strip()
        return existing["name"]
    master["buckets"][key] = {"name": display_name, "email": str(email).strip()}
    return display_name


def canonical_bucket_name(master: dict[str, Any], bucket_name: Any) -> str:
    display_name = str(bucket_name).strip() if bucket_name is not None else ""
    if not display_name:
        return ""
    key = normalize_lookup_key(display_name)
    existing = master["buckets"].get(key)
    if existing:
        return existing["name"]
    return add_bucket(master, display_name, "")


def get_bucket_email(master: dict[str, Any], bucket_name: Any) -> str:
    key = normalize_lookup_key(bucket_name)
    if not key:
        return ""
    bucket = master["buckets"].get(key)
    return bucket["email"] if bucket else ""


def ordered_union(values: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        ordered.append(value)
    return ordered


def get_program_row_key(program_no: Any, abbr: Any, row_number: int | None = None) -> str:
    program_key = normalize_lookup_key(program_no)
    if program_key:
        return f"PROG::{program_key}"
    abbr_key = normalize_lookup_key(abbr)
    if abbr_key:
        return f"ABBR::{abbr_key}"
    if row_number is not None:
        return f"ROW::{row_number}"
    return "ROW::UNKNOWN"


def rebuild_master_indexes(master: dict[str, Any]) -> None:
    degree_index: dict[str, dict[str, Any]] = {}
    programs_by_prog: dict[str, dict[str, Any]] = {}
    programs_by_abbr: dict[str, dict[str, Any]] = {}

    for degree_key, row in master["degree_rows"].items():
        degree_index[degree_key] = row
        bucket_name = canonical_bucket_name(master, row.get("MERGE_NAME"))
        row["MERGE_NAME"] = bucket_name
        row["EMAIL"] = get_bucket_email(master, bucket_name) if bucket_name else ""

    for row_key, row in master["program_rows"].items():
        bucket_name = canonical_bucket_name(master, row.get("MERGE_NAME"))
        row["MERGE_NAME"] = bucket_name
        row["EMAIL"] = get_bucket_email(master, bucket_name) if bucket_name else ""

        program_key = normalize_lookup_key(row.get("PROG_NO"))
        abbr_key = normalize_lookup_key(row.get("ABBR"))
        if program_key:
            programs_by_prog[program_key] = row
        if abbr_key:
            programs_by_abbr[abbr_key] = row

    master["indexes"] = {
        "degrees": degree_index,
        "programs_by_prog": programs_by_prog,
        "programs_by_abbr": programs_by_abbr,
    }


def read_sheet_dicts(worksheet) -> list[dict[str, Any]]:
    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        return []
    headers = [normalize_header(value) for value in header_row]
    records: list[dict[str, Any]] = []
    for row in rows:
        if row is None or is_blank_row(row):
            continue
        record: dict[str, Any] = {}
        for index in range(min(len(headers), len(row))):
            header = headers[index]
            if not header:
                continue
            record[header] = row[index]
        if record:
            records.append(record)
    return records


def load_bucket_master(bucket_master_path: Path) -> dict[str, Any]:
    if not bucket_master_path.exists():
        return create_empty_master()

    workbook = load_workbook(bucket_master_path, read_only=True, data_only=True)
    master = create_empty_master()
    try:
        if SHEET_BUCKETS in workbook.sheetnames:
            for row in read_sheet_dicts(workbook[SHEET_BUCKETS]):
                add_bucket(master, row.get("MERGE_NAME"), row.get("EMAIL"))

        if SHEET_DEGREES in workbook.sheetnames:
            for row in read_sheet_dicts(workbook[SHEET_DEGREES]):
                degree_name = str(row.get("DEGNM") or "").strip()
                if not degree_name:
                    continue
                bucket_name = canonical_bucket_name(master, row.get("MERGE_NAME"))
                degree_key = normalize_lookup_key(degree_name)
                master["degree_rows"][degree_key] = {
                    "DEGNM": degree_name,
                    "MERGE_NAME": bucket_name,
                    "EMAIL": get_bucket_email(master, bucket_name) if bucket_name else "",
                    "SOURCE": str(row.get("SOURCE") or "").strip(),
                    "STATUS": str(row.get("STATUS") or "").strip(),
                    "PROGRAM_COUNT": row.get("PROGRAM_COUNT"),
                    "FACULTIES": str(row.get("FACULTIES") or "").strip(),
                }

        if SHEET_PROGRAMS in workbook.sheetnames:
            for row_number, row in enumerate(
                read_sheet_dicts(workbook[SHEET_PROGRAMS]), start=2
            ):
                bucket_name = canonical_bucket_name(master, row.get("MERGE_NAME"))
                row_key = get_program_row_key(
                    row.get("PROG_NO"),
                    row.get("ABBR"),
                    row_number=row_number,
                )
                master["program_rows"][row_key] = {
                    "PROG_NO": str(row.get("PROG_NO") or "").strip(),
                    "ABBR": str(row.get("ABBR") or "").strip(),
                    "FACULTY": str(row.get("FACULTY") or "").strip(),
                    "DEGNM": str(row.get("DEGNM") or "").strip(),
                    "MDEGNM": str(row.get("MDEGNM") or "").strip(),
                    "SUBDEGNM": str(row.get("SUBDEGNM") or "").strip(),
                    "MSUBDEGNM": str(row.get("MSUBDEGNM") or "").strip(),
                    "MERGE_NAME": bucket_name,
                    "EMAIL": get_bucket_email(master, bucket_name) if bucket_name else "",
                    "SOURCE": str(row.get("SOURCE") or "").strip(),
                    "STATUS": str(row.get("STATUS") or "").strip(),
                }
    finally:
        workbook.close()

    rebuild_master_indexes(master)
    return master


def save_bucket_master(bucket_master_path: Path, master: dict[str, Any]) -> None:
    bucket_master_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    notes = workbook.active
    notes.title = SHEET_NOTES
    notes.append(["What to edit"])
    notes.append(
        [
            "Edit MERGE_NAME in Degree_List for whole-degree defaults. Edit MERGE_NAME in Program_Mapping for program-specific overrides."
        ]
    )
    notes.append(
        [
            "The merge script reads this Excel file on every run. Your edits here are used for future separation."
        ]
    )
    notes.append(
        [
            "Bucket_List controls the available bucket names and emails shown in the GUI."
        ]
    )

    bucket_sheet = workbook.create_sheet(SHEET_BUCKETS)
    bucket_sheet.append(["MERGE_NAME", "EMAIL"])
    for bucket in master["buckets"].values():
        bucket_sheet.append([bucket["name"], bucket["email"]])

    degree_sheet = workbook.create_sheet(SHEET_DEGREES)
    degree_sheet.append(DEGREE_MASTER_COLUMNS)
    for row in master["degree_rows"].values():
        degree_sheet.append(
            [
                row.get("DEGNM", ""),
                row.get("MERGE_NAME", ""),
                row.get("EMAIL", ""),
                row.get("SOURCE", ""),
                row.get("STATUS", ""),
                row.get("PROGRAM_COUNT", ""),
                row.get("FACULTIES", ""),
            ]
        )

    program_sheet = workbook.create_sheet(SHEET_PROGRAMS)
    program_sheet.append(PROGRAM_MASTER_COLUMNS)
    for row in master["program_rows"].values():
        program_sheet.append(
            [
                row.get("PROG_NO", ""),
                row.get("ABBR", ""),
                row.get("FACULTY", ""),
                row.get("DEGNM", ""),
                row.get("MDEGNM", ""),
                row.get("SUBDEGNM", ""),
                row.get("MSUBDEGNM", ""),
                row.get("MERGE_NAME", ""),
                row.get("EMAIL", ""),
                row.get("SOURCE", ""),
                row.get("STATUS", ""),
            ]
        )

    workbook.save(bucket_master_path)


def read_program_master_records(program_master_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(program_master_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return []

        headers = [normalize_header(value) for value in header_row]
        records: list[dict[str, str]] = []
        for row in rows:
            if row is None or is_blank_row(row):
                continue
            record: dict[str, str] = {}
            for index in range(min(len(headers), len(row))):
                header = headers[index]
                if not header:
                    continue
                record[header] = "" if row[index] is None else str(row[index]).strip()
            if record:
                records.append(record)
        return records
    finally:
        workbook.close()


def resolve_row_from_indexes(
    master: dict[str, Any],
    program_no: Any = None,
    abbr: Any = None,
    degree_name: Any = None,
) -> tuple[str, dict[str, Any] | None]:
    abbr_key = normalize_lookup_key(abbr)
    if abbr_key:
        row = master["indexes"]["programs_by_abbr"].get(abbr_key)
        if row and str(row.get("MERGE_NAME") or "").strip():
            return row["MERGE_NAME"], row

    program_key = normalize_lookup_key(program_no)
    if program_key:
        row = master["indexes"]["programs_by_prog"].get(program_key)
        if row and str(row.get("MERGE_NAME") or "").strip():
            return row["MERGE_NAME"], row

    degree_key = normalize_lookup_key(degree_name)
    if degree_key:
        row = master["indexes"]["degrees"].get(degree_key)
        if row and str(row.get("MERGE_NAME") or "").strip():
            return row["MERGE_NAME"], row

    return "", None


def sync_master_from_program_master(
    master: dict[str, Any], program_master_path: Path
) -> None:
    records = read_program_master_records(program_master_path)
    existing_degree_rows = master["degree_rows"]
    existing_program_rows = master["program_rows"]
    updated_program_rows: OrderedDict[str, dict[str, Any]] = OrderedDict()
    degree_stats: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"count": 0, "faculties": set(), "sample_degree": ""}
    )

    for record in records:
        degree_name = str(record.get("DEGNM") or "").strip()
        degree_key = normalize_lookup_key(degree_name)
        program_no = record.get("PROG_NO")
        abbr = record.get("ABBR")
        row_key = get_program_row_key(program_no, abbr)
        existing_program = existing_program_rows.get(row_key)
        existing_degree = existing_degree_rows.get(degree_key)

        bucket_name = ""
        source = ""
        if existing_program and str(existing_program.get("MERGE_NAME") or "").strip():
            bucket_name = canonical_bucket_name(master, existing_program.get("MERGE_NAME"))
            source = str(existing_program.get("SOURCE") or "").strip() or "EXCEL:PROGRAM"
        elif existing_degree and str(existing_degree.get("MERGE_NAME") or "").strip():
            bucket_name = canonical_bucket_name(master, existing_degree.get("MERGE_NAME"))
            source = str(existing_degree.get("SOURCE") or "").strip() or "EXCEL:DEGREE"
        else:
            bucket_name, source = classify_degree_name(degree_name)
            bucket_name = canonical_bucket_name(master, bucket_name)

        status = "MAPPED" if bucket_name else "REVIEW_NEEDED"
        updated_program_rows[row_key] = {
            "PROG_NO": str(program_no or "").strip(),
            "ABBR": str(abbr or "").strip(),
            "FACULTY": str(record.get("FACULTY") or "").strip(),
            "DEGNM": degree_name,
            "MDEGNM": str(record.get("MDEGNM") or "").strip(),
            "SUBDEGNM": str(record.get("SUBDEGNM") or "").strip(),
            "MSUBDEGNM": str(record.get("MSUBDEGNM") or "").strip(),
            "MERGE_NAME": bucket_name,
            "EMAIL": get_bucket_email(master, bucket_name) if bucket_name else "",
            "SOURCE": source,
            "STATUS": status,
        }

        degree_stats[degree_key]["count"] += 1
        faculty = str(record.get("FACULTY") or "").strip()
        if faculty:
            degree_stats[degree_key]["faculties"].add(faculty)
        degree_stats[degree_key]["sample_degree"] = degree_name

    for row_key, existing_row in existing_program_rows.items():
        if row_key in updated_program_rows:
            continue
        bucket_name = canonical_bucket_name(master, existing_row.get("MERGE_NAME"))
        updated_program_rows[row_key] = {
            **existing_row,
            "MERGE_NAME": bucket_name,
            "EMAIL": get_bucket_email(master, bucket_name) if bucket_name else "",
            "STATUS": existing_row.get("STATUS") or ("MAPPED" if bucket_name else "REVIEW_NEEDED"),
        }
        degree_name = str(existing_row.get("DEGNM") or "").strip()
        degree_key = normalize_lookup_key(degree_name)
        if degree_name:
            degree_stats[degree_key]["sample_degree"] = degree_name

    master["program_rows"] = updated_program_rows

    degree_keys = ordered_union(
        list(degree_stats.keys()) + list(existing_degree_rows.keys())
    )
    updated_degree_rows: OrderedDict[str, dict[str, Any]] = OrderedDict()

    for degree_key in degree_keys:
        stats = degree_stats.get(
            degree_key, {"count": 0, "faculties": set(), "sample_degree": ""}
        )
        existing_degree = existing_degree_rows.get(degree_key)
        degree_name = (
            str(existing_degree.get("DEGNM") or "").strip()
            if existing_degree
            else stats["sample_degree"]
        )
        bucket_name = ""
        source = ""
        if existing_degree and str(existing_degree.get("MERGE_NAME") or "").strip():
            bucket_name = canonical_bucket_name(master, existing_degree.get("MERGE_NAME"))
            source = str(existing_degree.get("SOURCE") or "").strip() or "EXCEL:DEGREE"
        else:
            bucket_name, source = classify_degree_name(degree_name)
            bucket_name = canonical_bucket_name(master, bucket_name)

        status = "MAPPED" if bucket_name else "REVIEW_NEEDED"
        updated_degree_rows[degree_key] = {
            "DEGNM": degree_name,
            "MERGE_NAME": bucket_name,
            "EMAIL": get_bucket_email(master, bucket_name) if bucket_name else "",
            "SOURCE": source,
            "STATUS": status,
            "PROGRAM_COUNT": stats["count"],
            "FACULTIES": ", ".join(sorted(stats["faculties"])),
        }

    master["degree_rows"] = updated_degree_rows
    master["program_master_path"] = str(program_master_path.resolve())
    rebuild_master_indexes(master)


def detect_program_master(
    selected_files: list[Path], script_dir: Path, cli_path: str | None
) -> Path | None:
    candidates: list[Path] = []
    if cli_path:
        candidates.append(Path(cli_path))
    for base_dir in [script_dir, *{path.parent for path in selected_files}]:
        candidates.append(base_dir / "program_master.xlsx")

    seen: set[Path] = set()
    for candidate in candidates:
        resolved = candidate.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        if resolved.exists():
            return resolved
    return None


def default_bucket_master_path(script_dir: Path, args: argparse.Namespace) -> Path:
    if args.bucket_master:
        return Path(args.bucket_master).resolve()
    return (script_dir / DEFAULT_BUCKET_MASTER).resolve()


def select_paths_with_gui(
    detected_program_master: Path | None, default_input_dir: Path
) -> tuple[list[Path], Path | None, Path | None]:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    input_files = filedialog.askopenfilenames(
        title="Select Excel files to merge",
        initialdir=str(default_input_dir),
        filetypes=[("Excel files", "*.xlsx")],
    )
    selected_files = [Path(path).resolve() for path in input_files]
    if not selected_files:
        messagebox.showinfo("Merge Cancelled", "No input files were selected.")
        return [], None, None

    program_master_path = detected_program_master
    if program_master_path is None:
        selected = filedialog.askopenfilename(
            title="Select Program Master Excel",
            initialdir=str(selected_files[0].parent),
            filetypes=[("Excel files", "*.xlsx")],
        )
        program_master_path = Path(selected).resolve() if selected else None

    output_folder = filedialog.askdirectory(
        title="Select output folder",
        initialdir=str(selected_files[0].parent),
    )
    output_dir = Path(output_folder).resolve() if output_folder else None
    return selected_files, program_master_path, output_dir


def list_input_files_from_cli(args: argparse.Namespace, script_dir: Path) -> list[Path]:
    selected_files: list[Path] = []
    if args.input_files:
        selected_files.extend(Path(path).resolve() for path in args.input_files)
    elif args.input_dir:
        input_dir = Path(args.input_dir).resolve()
        selected_files.extend(
            path.resolve()
            for path in sorted(input_dir.glob("*.xlsx"))
            if not path.name.startswith("~$")
        )
    elif args.no_gui:
        input_dir = script_dir.resolve()
        selected_files.extend(
            path.resolve()
            for path in sorted(input_dir.glob("*.xlsx"))
            if not path.name.startswith("~$")
        )
    return selected_files


def clean_selected_input_files(
    selected_files: list[Path],
    program_master_path: Path | None,
    bucket_master_path: Path,
) -> list[Path]:
    cleaned: list[Path] = []
    for path in selected_files:
        if path.suffix.lower() != ".xlsx":
            continue
        if path.name.startswith("~$"):
            continue
        if program_master_path and path.resolve() == program_master_path.resolve():
            continue
        if path.resolve() == bucket_master_path.resolve():
            continue
        cleaned.append(path.resolve())
    return cleaned


def prompt_user_for_bucket(
    master: dict[str, Any],
    degree_name: Any,
    program_no: Any,
    abbr: Any,
    file_name: str,
    gui_enabled: bool,
) -> str | None:
    if not gui_enabled:
        raise RuntimeError(
            f'No saved mapping found for DEGNM "{degree_name}" in file "{file_name}". '
            "Run the script without --no-gui so you can choose a bucket once in the GUI."
        )

    import tkinter as tk

    root = tk._default_root
    if root is None:
        root = tk.Tk()
        root.withdraw()

    result = {"bucket": None}
    window = tk.Toplevel(root)
    window.title("Select Merge Bucket")
    window.resizable(False, False)

    lines = [
        "This row could not be mapped from the Excel master.",
        f"DEGNM: {degree_name or '(blank)'}",
        f"Program: {program_no or '(blank)'}",
        f"ABBR: {abbr or '(blank)'}",
        f"Source File: {file_name}",
        "",
        "Select the bucket to save for future runs.",
    ]

    label = tk.Label(window, text="\n".join(lines), justify="left", padx=16, pady=16)
    label.pack(fill="both")

    button_frame = tk.Frame(window, padx=16, pady=8)
    button_frame.pack(fill="both")

    def choose(bucket_name: str) -> None:
        result["bucket"] = bucket_name
        window.destroy()

    for bucket in master["buckets"].values():
        button = tk.Button(
            button_frame,
            text=f"{bucket['name']}  |  {bucket['email']}",
            width=48,
            command=lambda choice=bucket["name"]: choose(choice),
        )
        button.pack(fill="x", pady=4)

    cancel_button = tk.Button(button_frame, text="Cancel", command=window.destroy)
    cancel_button.pack(fill="x", pady=(8, 0))

    window.grab_set()
    window.focus_force()
    window.wait_window()
    return result["bucket"]


def can_promote_to_degree_default(
    master: dict[str, Any], degree_key: str, chosen_bucket: str
) -> bool:
    existing_degree = master["degree_rows"].get(degree_key)
    existing_bucket = (
        canonical_bucket_name(master, existing_degree.get("MERGE_NAME"))
        if existing_degree
        else ""
    )
    if existing_bucket and existing_bucket != chosen_bucket:
        return False

    mapped_buckets = set()
    for row in master["program_rows"].values():
        if normalize_lookup_key(row.get("DEGNM")) != degree_key:
            continue
        bucket_name = canonical_bucket_name(master, row.get("MERGE_NAME"))
        if bucket_name:
            mapped_buckets.add(bucket_name)

    if not mapped_buckets:
        return True
    return mapped_buckets == {chosen_bucket}


def remember_manual_selection(
    master: dict[str, Any],
    chosen_bucket: str,
    row_dict: dict[str, Any],
) -> None:
    bucket_name = canonical_bucket_name(master, chosen_bucket)
    program_no = row_dict.get("APPL_NO") or row_dict.get("PROG_NO")
    abbr = row_dict.get("ABBR")

    matched_row = None
    abbr_key = normalize_lookup_key(abbr)
    if abbr_key:
        matched_row = master["indexes"]["programs_by_abbr"].get(abbr_key)

    if matched_row is None:
        program_key = normalize_lookup_key(program_no)
        if program_key:
            matched_row = master["indexes"]["programs_by_prog"].get(program_key)

    if matched_row is None:
        row_key = get_program_row_key(program_no, abbr, row_number=len(master["program_rows"]) + 2)
        matched_row = {
            "PROG_NO": str(program_no or "").strip(),
            "ABBR": str(abbr or "").strip(),
            "FACULTY": str(row_dict.get("Faculty") or row_dict.get("FACULTY") or "").strip(),
            "DEGNM": str(row_dict.get("DEGNM") or "").strip(),
            "MDEGNM": str(row_dict.get("MDEGNM") or "").strip(),
            "SUBDEGNM": str(row_dict.get("SUBDEGNM") or "").strip(),
            "MSUBDEGNM": str(row_dict.get("MSUBDEGNM") or "").strip(),
            "MERGE_NAME": "",
            "EMAIL": "",
            "SOURCE": "",
            "STATUS": "",
        }
        master["program_rows"][row_key] = matched_row

    matched_row["MERGE_NAME"] = bucket_name
    matched_row["EMAIL"] = get_bucket_email(master, bucket_name)
    matched_row["SOURCE"] = "MANUAL:GUI_SELECTION"
    matched_row["STATUS"] = "MAPPED"

    degree_name = str(row_dict.get("DEGNM") or "").strip()
    degree_key = normalize_lookup_key(degree_name)
    if degree_name and can_promote_to_degree_default(master, degree_key, bucket_name):
        degree_row = master["degree_rows"].get(degree_key)
        if degree_row is None:
            degree_row = {
                "DEGNM": degree_name,
                "MERGE_NAME": "",
                "EMAIL": "",
                "SOURCE": "",
                "STATUS": "",
                "PROGRAM_COUNT": "",
                "FACULTIES": "",
            }
            master["degree_rows"][degree_key] = degree_row
        degree_row["MERGE_NAME"] = bucket_name
        degree_row["EMAIL"] = get_bucket_email(master, bucket_name)
        degree_row["SOURCE"] = "MANUAL:GUI_SELECTION"
        degree_row["STATUS"] = "MAPPED"

    rebuild_master_indexes(master)


def resolve_row_bucket(
    row_dict: dict[str, Any],
    master: dict[str, Any],
    gui_enabled: bool,
    source_file_name: str,
) -> str:
    bucket_name, _ = resolve_row_from_indexes(
        master,
        program_no=row_dict.get("APPL_NO") or row_dict.get("PROG_NO"),
        abbr=row_dict.get("ABBR"),
        degree_name=row_dict.get("DEGNM"),
    )
    if bucket_name:
        return bucket_name

    chosen_bucket = prompt_user_for_bucket(
        master=master,
        degree_name=row_dict.get("DEGNM"),
        program_no=row_dict.get("APPL_NO") or row_dict.get("PROG_NO"),
        abbr=row_dict.get("ABBR"),
        file_name=source_file_name,
        gui_enabled=gui_enabled,
    )
    if not chosen_bucket:
        raise RuntimeError(
            f'Merge was cancelled because no bucket was selected for DEGNM "{row_dict.get("DEGNM")}".'
        )

    remember_manual_selection(master, chosen_bucket, row_dict)
    return chosen_bucket


def collect_rows_by_bucket(
    input_files: list[Path],
    sheet_name: str,
    master: dict[str, Any],
    gui_enabled: bool,
) -> tuple[dict[str, list[dict[str, Any]]], list[str]]:
    rows_by_bucket: dict[str, list[dict[str, Any]]] = defaultdict(list)
    discovered_columns: set[str] = set()
    extra_columns: list[str] = []

    for file_path in input_files:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            data_sheet_name = get_default_sheet_name(workbook, sheet_name)
            worksheet = workbook[data_sheet_name]
            iterator = worksheet.iter_rows(values_only=True)
            header_row = next(iterator, None)
            if header_row is None:
                continue

            headers = [normalize_header(value) for value in header_row]
            if "DEGNM" not in headers:
                raise ValueError(f'Column "DEGNM" not found in file "{file_path.name}".')

            for header in headers:
                if not header:
                    continue
                discovered_columns.add(header)
                if header not in STANDARD_COLUMNS and header not in extra_columns:
                    extra_columns.append(header)

            for row in iterator:
                if row is None or is_blank_row(row):
                    continue

                row_dict = {
                    headers[index]: row[index]
                    for index in range(min(len(headers), len(row)))
                    if headers[index]
                }
                bucket_name = resolve_row_bucket(
                    row_dict=row_dict,
                    master=master,
                    gui_enabled=gui_enabled,
                    source_file_name=file_path.name,
                )
                rows_by_bucket[bucket_name].append(row_dict)
        finally:
            workbook.close()

    column_order = [column for column in STANDARD_COLUMNS if column in discovered_columns]
    column_order.extend(extra_columns)
    return rows_by_bucket, column_order


def write_bucket_workbook(
    output_path: Path,
    sheet_name: str,
    column_order: list[str],
    rows: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(column_order)

    for row_index, row_dict in enumerate(rows, start=2):
        for col_index, column_name in enumerate(column_order, start=1):
            value = row_dict.get(column_name)
            cell = worksheet.cell(row=row_index, column=col_index)

            if column_name in TEXT_ONLY_COLUMNS:
                cell.value = as_text(value)
                cell.number_format = "@"
                continue

            if column_name == CLASS_COLUMN:
                numeric_value = parse_numeric_class(value)
                if numeric_value is not None:
                    cell.value = numeric_value
                    cell.number_format = "0.00"
                else:
                    cell.value = value
                    if value is not None:
                        cell.number_format = "@"
                continue

            cell.value = value

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def cleanup_old_output_files(output_dir: Path, bucket_names: list[str]) -> None:
    for bucket_name in bucket_names:
        path = output_dir / f"{safe_filename(bucket_name)}.xlsx"
        if path.exists():
            path.unlink()


def write_output_files(
    output_dir: Path,
    sheet_name: str,
    rows_by_bucket: dict[str, list[dict[str, Any]]],
    column_order: list[str],
    master: dict[str, Any],
) -> dict[str, Path]:
    output_paths: dict[str, Path] = {}
    cleanup_old_output_files(
        output_dir, [bucket["name"] for bucket in master["buckets"].values()]
    )
    for bucket_name, rows in rows_by_bucket.items():
        if not rows:
            continue
        output_path = output_dir / f"{safe_filename(bucket_name)}.xlsx"
        write_bucket_workbook(output_path, sheet_name, column_order, rows)
        output_paths[bucket_name] = output_path
    return output_paths


def show_success_message(
    master: dict[str, Any],
    output_paths: dict[str, Path],
    rows_by_bucket: dict[str, list[dict[str, Any]]],
    gui_enabled: bool,
) -> None:
    lines = []
    for bucket in master["buckets"].values():
        bucket_name = bucket["name"]
        if bucket_name in output_paths:
            lines.append(
                f"{bucket_name}: {len(rows_by_bucket[bucket_name])} rows -> {output_paths[bucket_name]}"
            )
    message = "\n".join(lines)
    print(message)

    if gui_enabled:
        import tkinter as tk
        from tkinter import messagebox

        root = tk._default_root
        if root is None:
            root = tk.Tk()
            root.withdraw()
        messagebox.showinfo("Merge Completed", message or "No output files were created.")


def main() -> None:
    args = parse_args()
    script_dir = Path(__file__).resolve().parent
    bucket_master_path = default_bucket_master_path(script_dir, args)
    master = load_bucket_master(bucket_master_path)

    selected_files = list_input_files_from_cli(args, script_dir)
    detected_program_master = detect_program_master(
        selected_files=selected_files,
        script_dir=script_dir,
        cli_path=args.program_master,
    )

    gui_enabled = not args.no_gui
    if gui_enabled:
        default_input_dir = selected_files[0].parent if selected_files else script_dir
        selected_files, selected_program_master, output_dir = select_paths_with_gui(
            detected_program_master=detected_program_master,
            default_input_dir=default_input_dir,
        )
        if not selected_files:
            return
        program_master_path = selected_program_master
    else:
        output_dir = (
            Path(args.output_dir).resolve()
            if args.output_dir
            else script_dir / "merged_by_degree"
        )
        program_master_path = detected_program_master

    if program_master_path is None and not bucket_master_path.exists():
        raise SystemExit(
            "Program master workbook could not be found. Please select or pass --program-master."
        )

    if program_master_path is not None:
        if not program_master_path.exists():
            raise SystemExit(f"Program master workbook not found: {program_master_path}")
        sync_master_from_program_master(master, program_master_path)
        save_bucket_master(bucket_master_path, master)
    else:
        rebuild_master_indexes(master)

    if args.build_master_only:
        print(f"Saved editable bucket master: {bucket_master_path}")
        return

    cleaned_input_files = clean_selected_input_files(
        selected_files=selected_files,
        program_master_path=program_master_path,
        bucket_master_path=bucket_master_path,
    )
    if not cleaned_input_files:
        raise SystemExit(
            "No source Excel files were selected after excluding the master workbooks."
        )

    output_dir = output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    rows_by_bucket, column_order = collect_rows_by_bucket(
        input_files=cleaned_input_files,
        sheet_name=args.sheet_name,
        master=master,
        gui_enabled=gui_enabled,
    )
    save_bucket_master(bucket_master_path, master)

    if not rows_by_bucket:
        raise SystemExit("No data rows were found to merge.")

    output_paths = write_output_files(
        output_dir=output_dir,
        sheet_name=args.sheet_name,
        rows_by_bucket=rows_by_bucket,
        column_order=column_order,
        master=master,
    )
    show_success_message(master, output_paths, rows_by_bucket, gui_enabled)


if __name__ == "__main__":
    main()
